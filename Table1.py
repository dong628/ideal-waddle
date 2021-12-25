from openpyxl import *
import datetime as d
import time as t
import sys

## Default Settings

ColDelta = 1; RowDelta = 2;
BdDelta = 4;
BdAlpha = [1, 2, 4, 5, 6, 8, 9];
Filename = input("Filename of Vanilla:")
# LastS = int(input("LastStart:"))
# LastE = int(input("LastEnd:"))
CurDay = int(input("CurDay:"))-1
CurS = d.date.today()+d.timedelta(days=1);
CurE = d.date.today()+d.timedelta(days=1+CurDay);
DatePointV = "A1"; DatePointF = "A9";
DatePointB = "E2";
DateFormatV = "TG2102座位表 （%s—%s）"
DateFormatF = "讲台\nTG2102座位表 （%s—%s）"
DateFormatB = "有效日期：%s-%s"
FlipAvailable = False; BdAvailable = False;
ValidDate = (CurS.strftime("%Y.%m.%d"), CurE.strftime("%Y.%m.%d"));
# CurS = int(CurS.strftime("%y%m%d"))
# CurE = int(CurE.strftime("%y%m%d"))

## Vallina part

try:
	Vanilla = load_workbook("Vanilla_%s.xlsx" % Filename);
except:
	print("Vanilla Filename Error!")
	sys.exit()
# CurrentTime = t.strftime("%y%m", t.localtime());
# StartDay = int(t.strftime("%d", t.localtime()))-int(t.strftime("%w", t.localtime())))+1
# if int(t.strftime("%w", t.localtime()))==0:

vanilla = Vanilla["Sheet1"]

'''
male = list()
female = list()

for i in range(ColDelta+1, ColDelta+6, 2):
	male.append(chr(i+65));

for i in range(ColDelta, ColDelta+7, 2):
	female.append(chr(i+65));
'''

for i in reversed(list(range(ColDelta, ColDelta+7))):
	for j in range(RowDelta+1, RowDelta+8):
		point1 = chr(i+65)+str(j);
		point2 = chr(i+65+2)+str(j+1);
		vanilla[point2].value = vanilla[point1].value; 
		vanilla[point1].value = str();

pointv = chr(ColDelta+65+7);
for i in range(RowDelta+2, RowDelta+9):
	point1 = pointv + str(i);
	point2 = chr(ColDelta+66) + str(i);
	vanilla[point2].value = vanilla[point1].value; 
	vanilla[point1].value = str();

pointv = chr(ColDelta+65+8);
for i in range(RowDelta+2, RowDelta+9):
	point1 = pointv + str(i);
	point2 = chr(ColDelta+65) + str(i);
	vanilla[point2].value = vanilla[point1].value; 
	vanilla[point1].value = str();

for i in range(ColDelta, ColDelta+7):
	point1 = chr(i+65) + str(RowDelta+8);
	point2 = chr(i+65) + str(RowDelta+1);
	vanilla[point2].value = vanilla[point1].value; 
	vanilla[point1].value = str();

vanilla[DatePointV].value = DateFormatV % ValidDate;

## Flip part

try:
	Flip = load_workbook("Flip_%s.xlsx" % Filename);
except:
	print("Flip Filename Error!");
else:
	flip = Flip["Sheet1"];
	FlipAvailable = True;

	for i in range(7):
		for j in range(7):
			point1 = chr(j+ColDelta+65) + str(i+RowDelta+1);
			point2 = chr(6-j+ColDelta+65) + str(6-i+1);
			flip[point2].value = vanilla[point1].value;

flip[DatePointF].value = DateFormatF % ValidDate;

## BackDoor part

try:
	BackDoor = load_workbook("BackDoor_%s.xlsx" % Filename);
except:
	print("BackDoor Filename Error!");
else:
	bd = BackDoor["TG2102"];
	BdAvailable = True;

	bd.row_dimensions[4].height = 26 # Temp
	for i in range(7):
		Row = bd.row_dimensions[i+BdDelta+1]; # Temp
		Row.height = 26; # Temp
		for j in range(7):
			point1 = chr(j+ColDelta+65) + str(i+RowDelta+1);
			point2 = chr(65+BdAlpha[j]) + str(i+BdDelta+1);
			bd[point2].value = vanilla[point1].value;

bd[DatePointB].value = DateFormatB % ValidDate;

## Save workbooks

if FlipAvailable:
	Flip.save("Flip_%s-%s.xlsx" % (CurS.strftime("%y%m%d"), CurE.strftime("%y%m%d")));
if BdAvailable:
	BackDoor.save("BackDoor_%s-%s.xlsx" % (CurS.strftime("%y%m%d"), CurE.strftime("%y%m%d")));
Vanilla.save("Vanilla_%s-%s.xlsx" % (CurS.strftime("%y%m%d"), CurE.strftime("%y%m%d")));
